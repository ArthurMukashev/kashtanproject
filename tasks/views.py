from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.http import HttpResponse
from django.shortcuts import render, redirect

from tasks.models import Variant, Task, UserTask


# Create your views here.

def index(request):
    variants = Variant.objects.all()
    return render(request, 'index.html', {'variants': variants})


def glossary(request):
    return render(request, 'glossary.html')


@login_required(login_url='/accounts/login/')
def profile(request):
    if request.user.is_superuser:
        group = "Администраторы"
    else:
        group = str(Group.objects.get(user=request.user))

    context_dict = {'group': group}

    if group == "Преподаватели" or group == "Администраторы":
        user_tasks = UserTask.objects.filter(status=0)
        context_dict['user_tasks'] = user_tasks

    if group == "Ученики":
        user_tasks = UserTask.objects.filter(user=request.user)
        context_dict['user_tasks'] = user_tasks

    return render(request, 'profile.html', context_dict)


@login_required(login_url='/accounts/login')
def variant(request, pk):
    if request.method == 'POST':
        d = {
            key: val if len(val) > 1 else val[0]
            for key, val in request.POST.lists()
        }
        z = list(zip(d.get('num_tasks[]'), d.get('tasks[]')))

        for task in z:
            t = Task.objects.get(pk=task[0])
            user_task = UserTask(task=t, user=request.user, answer=task[1])
            user_task.save()
        return render(request, 'variant_success.html')
    variant_obj = Variant.objects.get(pk=pk)
    tasks = variant_obj.tasks.all()
    return render(request, 'variant.html', {'variant': variant_obj, 'tasks': tasks})


@login_required(login_url='/accounts/login')
def tasks_verify(request, task_id, price, user_id):
    user_task = UserTask.objects.filter(task=task_id, user=user_id)
    user_task.update(price=price, status=True, pricer=request.user)
    return redirect('tasks:profile')


@login_required(login_url='/accounts/login')
def excel(request):
    # UserTasks оттуда взять всех юзеров с ГРУППИРОВКОЙ
    # потом надо на каждого юзера сделать ТУДА ЖЕ запрос и достать его оценки, сортировать по таскам
    # и в зависимости от таска выводить строчки, а потом уже ИТОГО и подпись
    data = []

    queryset = UserTask.objects.order_by().values('user').distinct()
    for task in queryset:
        real_task_list = [0] * 19
        task_list = list(UserTask.objects.filter(user=task['user']))
        task_list.sort(key=lambda x: x.task.sort_index, reverse=True)
        for i in range(len(task_list)):
            real_task_list[i] = task_list[i].price
        user = User.objects.get(pk=task['user'])

        data.append({
            'user': user.first_name + " " + user.last_name,
            'tasks': real_task_list
        })

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb = load_workbook('report.xlsx')
    ws = wb['Лист1']
    ws.title = f'Отчет за {datetime.now().strftime("%Y-%m-%d")}'
    ft = Font(name='Times New Roman', size=12)
    brd = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                 bottom=Side(style='medium'))
    alg = Alignment(horizontal='center')

    row_num = 0
    student_num = 1
    columns = [
        ("№ п/п", 7),
        (u'Ученик', 30),
        ('№ 1', 5),
        ('№ 2', 5),
        ('№ 3', 5),
        ('№ 4', 5),
        ('№ 5', 5),
        ('№ 6', 5),
        ('№ 7', 5),
        ('№ 8', 5),
        ('№ 9', 5),
        ('№ 10', 5),
        ('№ 11', 5),
        ('№ 12', 5),
        ('№ 13', 5),
        ('№ 14', 5),
        ('№ 15', 5),
        ('№ 16', 5),
        ('№ 17', 5),
        ('№ 18', 5),
        ('№ 19', 5),
        ('Итого', 10)
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.font = ft
        c.border = brd
        c.alignment = alg
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for data_row in data:
        row_num += 1
        row = [
            student_num,
            data_row.get('user'),
        ]
        row += data_row.get('tasks')
        row.append(f'=SUM(C{row_num + 1}:U{row_num + 1})')
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.font = ft
            c.border = brd
            c.alignment = alg
            c.value = row[col_num]
        student_num += 1

    row_num += 3
    ws.merge_cells(f'T{row_num}:V{row_num}')

    c = ws.cell(row=row_num, column=20)
    c.font = ft
    c.alignment = Alignment(horizontal='right')
    c.value = datetime.now().strftime('%d.%m.%Y')

    c = ws.cell(row=row_num, column=2)
    c.font = ft
    c.value = request.user.first_name + " " + request.user.last_name

    wb.save(response)
    return response


@login_required(login_url='/accounts/login/')
def abort(request):
    UserTask.objects.all().delete()
    return redirect('tasks:profile')


@login_required(login_url='/accounts/login/')
def tasks_auto_check(request):
    tasks = UserTask.objects.filter(status=0)
    for task in tasks:
        if task.task.task_answer == task.answer:
            task_price = 1
        else:
            task_price = 0
        task.task_price = task_price
        task.status = True
        task.pricer = request.user
        task.save()
    return redirect('tasks:profile')


@login_required(login_url='/accounts/login/')
def tasks_argue(request, task_id):
    task = UserTask.objects.get(user_id=request.user.id, task_id=task_id)
    task.status = False
    task.save()
    return redirect('tasks:profile')
